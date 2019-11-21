import openpyxl
from datetime import datetime
import string, os

class JR1Report:
    """
    Represents the JR1 report spreadsheet.

    The JR1 report is defined by the COUNTER Code of Practice R4 specification.
    The layout and data requirements are very specific. The properties and methods
    defined in this class rely on adherence to this specification otherwise data
    anomalies may arise and/or code using this class may raise an exception.
    """

    def __init__(self, workbook):
        self._workbook = openpyxl.load_workbook(filename=workbook, data_only=True)
        self._worksheet = self._workbook.active
        self._period = self._worksheet.cell(row=5, column=1).value
        self._filename = os.path.basename(workbook)
    
    @property
    def filename(self):
        """
        Returns the Excel file name
        """
        
        return self._filename
    
    @property
    def platform(self):
        """
        Returns the platform name represented in cell C10
        """
        
        return self._worksheet.cell(row=10, column=3).value

    @property
    def period(self):
        """
        Returns the report period represented in cell A5
        """
        
        return self._period
    
    @property
    def period_from(self):
        """
        Returns the period from date represented in cell A5
        """
        
        return datetime.strptime(self._period.strip()[:10], '%Y-%m-%d')

    @property
    def period_to(self):
        """
        Returns the period to date represented in cell A5
        """
        
        return datetime.strptime(self._period.strip()[-10:], '%Y-%m-%d')

    def data_range(self):
        """
        Returns the range of rows containing publication data.

        According to the COUNTER Code of Practice R4, publication data always
        starts at row 10 in the spreadsheet.
        """

        row_num = 10
        for row in self._worksheet.iter_rows(min_row=10, min_col=1, max_row=10000, max_col=1):
            if row[0].value == None:
                break
            row_num = row_num + 1
            
        return list(range(10, row_num))
    
    def get_row(self, row_num):
        """
        Returns a complete row of publication data for the given row number.

        For a JR1 report covering twelve months of usage data (typically Jan-Dec),
        there will be 22 columns of data in the row. The first item in the list
        will always be the given row number. Reports containing less (or more)
        months of usage data will have less or more list items, respectively.
        """

        datarow = list()
        datarow.append(row_num)
        for row in self._worksheet.iter_cols(min_row=row_num, min_col=1, max_row=row_num, max_col=22):
            for cell in row:
                datarow.append(str(cell.value).replace('"', ''))
        
        return datarow
